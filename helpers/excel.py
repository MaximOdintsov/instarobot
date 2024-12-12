from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font


class PyXLWriter:
    def __init__(self, colors=3):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.colors = colors
        self.widths = {}

    def set_cell(self, pos, value, width=None, height=None):
        row, col = pos
        colname = get_column_letter(col)
        slen = len(value or '') * 1
        if width is None:
            width = max(8, min(12 * 8, slen))
        if height is None:
            height = 15 * (1 + slen // width)
        if height > 15 * 50:
            height = 15 * 50
        old_h = self.ws.row_dimensions[row].height
        self.ws.row_dimensions[row].height = max(height or 25, old_h or 25)
        if col not in self.widths:
            old_w = self.ws.column_dimensions[colname].width or 8
            self.ws.column_dimensions[colname].width = max(width, old_w)
        self.ws.column_dimensions[colname].bestFit = True
        cell = self.ws.cell(row=row, column=col)
        cell.value = value
        if self.colors and row >= self.colors:
            if (row - self.colors) % 2:  # e.g. 2, 4, 6, ...
                cell.fill = PatternFill(start_color='FFFFDD', end_color='FFFFDD', fill_type='solid')
            else:  # e.g. 3, 5, 7, ...
                cell.fill = PatternFill(start_color='DDFFFF', end_color='DDFFFF', fill_type='solid')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        return cell

    def add_borders(self, style: str = 'dashed', color: str = '000000'):
        """ Добавление линий между всеми заполненными ячейками таблицы """
        border_style = Side(style=style, color=color)
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

        # Определяем диапазон заполненных ячеек для обработки
        max_row = self.ws.max_row
        max_col = self.ws.max_column

        # Применяем границы ко всем заполненным ячейкам
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = self.ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.border = border

    def set_width(self, col, width):
        colname = get_column_letter(col)
        self.ws.column_dimensions[colname].width = width
        self.widths[col] = width

    def set_headers_style(self):
        """Стилизация хедеров"""
        for col in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(row=1, column=col)
            cell.font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
            cell.fill = PatternFill(start_color='452bd9', end_color='452bd9', fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
    def __setitem__(self, pos, value):
        self.set_cell(pos, str(value))

    def save(self, path):
        self.ws.calculate_dimension()
        self.set_headers_style()
        self.add_borders()
        
        self.wb.save(path)